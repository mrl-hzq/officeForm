from __future__ import annotations

import math
import shutil
import tempfile
from copy import copy
from datetime import date, datetime
from pathlib import Path

from scripts.libreoffice_export import run_calc_pdf_export


def _calculate_text_font_size(text: str, base_chars_per_line: int = 40, max_height: float = 85.0) -> float:
    if not text or not text.strip():
        return 11.0
    lines = text.splitlines()
    font_sizes = [11.0, 10.5, 10.0, 9.5, 9.0, 8.5, 8.0, 7.5, 7.0, 6.5, 6.0, 5.5, 5.0]
    for fs in font_sizes:
        chars_per_line = max(1, int(base_chars_per_line * (11.0 / fs)))
        total_lines = 0
        for line in lines:
            if not line:
                total_lines += 1
            else:
                total_lines += math.ceil(len(line) / chars_per_line)
        needed_height = total_lines * (fs * 1.30)
        if needed_height <= max_height:
            return fs
    return 5.0


def _format_kpi_feedback_cell(sheet, cell_address: str, text: str, max_height: float) -> None:
    import openpyxl
    if not text or not text.strip():
        sheet[cell_address].value = ""
        return
    cleaned = text.replace("\r\n", "\n").strip()
    fs = _calculate_text_font_size(cleaned, base_chars_per_line=85, max_height=max_height)
    cell = sheet[cell_address]
    cell.value = cleaned
    cell.font = openpyxl.styles.Font(name="Calibri", size=fs)
    cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top", horizontal="left")


def _unprotect_and_prepare_kpi_workbook(
    template_path: Path,
    working_workbook_path: Path,
    task_list: str,
    worker_feedback: str = "",
    training_needs: str = "",
    evaluator_feedback: str = "",
) -> None:
    working_workbook_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        import openpyxl
        wb = openpyxl.load_workbook(template_path)
        for sheet in wb.worksheets:
            sheet.protection.sheet = False

        sheet = wb["MASTER"]
        thin_side = openpyxl.styles.Side(border_style="thin", color="000000")
        no_side = openpyxl.styles.Side(border_style=None)

        if task_list and task_list.strip():
            cleaned = task_list.replace("\r\n", "\n").strip()
            blocks = [b.strip() for b in cleaned.split("\n\n") if b.strip()]

            if len(blocks) >= 2 or len(cleaned.splitlines()) >= 6:
                if len(blocks) < 2:
                    lines = [l.strip() for l in cleaned.splitlines() if l.strip()]
                    blocks = ["\n".join(lines[i:i + 2]) for i in range(0, len(lines), 2)]

                mid = (len(blocks) + 1) // 2
                left_text = "\n\n".join(blocks[:mid])
                right_text = "\n\n".join(blocks[mid:])

                fs_left = _calculate_text_font_size(left_text, base_chars_per_line=40)
                fs_right = _calculate_text_font_size(right_text, base_chars_per_line=38)
                chosen_fs = min(fs_left, fs_right)

                try:
                    sheet.unmerge_cells("A9:I10")
                except Exception:
                    pass
                sheet.merge_cells("A9:D10")
                sheet.merge_cells("E9:I10")

                for r in range(9, 11):
                    for c in range(1, 5):
                        l_b = thin_side if c == 1 else no_side
                        t_b = thin_side if r == 9 else no_side
                        r_b = no_side
                        b_b = thin_side if r == 10 else no_side
                        sheet.cell(row=r, column=c).border = openpyxl.styles.Border(left=l_b, top=t_b, right=r_b, bottom=b_b)

                for r in range(9, 11):
                    for c in range(5, 10):
                        l_b = no_side
                        t_b = thin_side if r == 9 else no_side
                        r_b = thin_side if c == 9 else no_side
                        b_b = thin_side if r == 10 else no_side
                        sheet.cell(row=r, column=c).border = openpyxl.styles.Border(left=l_b, top=t_b, right=r_b, bottom=b_b)

                cell_left = sheet["A9"]
                cell_left.value = left_text
                cell_left.font = openpyxl.styles.Font(name="Calibri", size=chosen_fs)
                cell_left.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top", horizontal="left")

                cell_right = sheet["E9"]
                cell_right.value = right_text
                cell_right.font = openpyxl.styles.Font(name="Calibri", size=chosen_fs)
                cell_right.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top", horizontal="left")
            else:
                for r in range(9, 11):
                    for c in range(1, 10):
                        l_b = thin_side if c == 1 else no_side
                        t_b = thin_side if r == 9 else no_side
                        r_b = thin_side if c == 9 else no_side
                        b_b = thin_side if r == 10 else no_side
                        sheet.cell(row=r, column=c).border = openpyxl.styles.Border(left=l_b, top=t_b, right=r_b, bottom=b_b)

                chosen_fs = _calculate_text_font_size(cleaned, base_chars_per_line=85)
                cell = sheet["A9"]
                cell.value = cleaned
                cell.font = openpyxl.styles.Font(name="Calibri", size=chosen_fs)
                cell.alignment = openpyxl.styles.Alignment(wrap_text=True, vertical="top", horizontal="left")
        else:
            for r in range(9, 11):
                for c in range(1, 10):
                    l_b = thin_side if c == 1 else no_side
                    t_b = thin_side if r == 9 else no_side
                    r_b = thin_side if c == 9 else no_side
                    b_b = thin_side if r == 10 else no_side
                    sheet.cell(row=r, column=c).border = openpyxl.styles.Border(left=l_b, top=t_b, right=r_b, bottom=b_b)
            sheet["A9"].value = ""

        _format_kpi_feedback_cell(sheet, "A106", worker_feedback, max_height=80.0)
        _format_kpi_feedback_cell(sheet, "A110", training_needs, max_height=50.0)
        _format_kpi_feedback_cell(sheet, "A117", evaluator_feedback, max_height=80.0)

        wb.save(working_workbook_path)
    except Exception:
        shutil.copyfile(template_path, working_workbook_path)

XL_TYPE_PDF = 0
XL_PAPER_A4 = 9
XL_LANDSCAPE = 2
XL_PRINTER = 2
XL_PICTURE = -4147
XL_CENTER = -4108
XL_RIGHT = -4152
XL_LEFT = -4131
MARGIN_POINTS = 18
PDF_SLOT_GUTTER_POINTS = 14
PDF_SLOT_PADDING_POINTS = 18
KPI_LAYOUT_SHEET_NAME = "_KPI_PDF_LAYOUT"
KPI_MASTER_PRINT_RANGES = ("A41:I68", "A69:I83", "A84:I124")
KPI_GUIDE_PRINT_RANGES = ("A1:E50", "A51:E88")

KPI_SCORE_CELLS = {
    "knowledge": ["H16", "H17", "H18", "H19", "H20"],
    "quality": ["H24", "H25", "H26", "H27", "H28"],
    "problemSolving": ["H32", "H33", "H34", "H35", "H36"],
    "communication": ["H40", "H41", "H42", "H43", "H44"],
    "teamwork": ["H48", "H49", "H50", "H51", "H52"],
    "initiative": ["H56", "H57", "H58", "H59", "H60"],
    "continuousLearning": ["H64", "H65", "H66", "H67", "H68"],
}

KPI_COMMENT_CELLS = {
    "knowledge": "B21",
    "quality": "B29",
    "problemSolving": "B37",
    "communication": "B45",
    "teamwork": "B53",
    "initiative": "B61",
    "continuousLearning": "B69",
}

KPI_OPTION_CELLS = {
    "breakfastMeeting": "F85",
    "emergencyLeaveAttendance": "F86",
    "medicalLeaveAttendance": "F87",
    "biroAgama": "F90",
    "biroSukan": "F91",
    "trainingHours": "F94",
    "committeeRole": "F95",
    "eqariah": "F98",
}


def _as_excel_date_text(value: str | date) -> str:
    if isinstance(value, date):
        parsed = value
    else:
        parsed = datetime.strptime(value, "%Y-%m-%d").date()
    return parsed.strftime("%d-%b-%Y")


def _set_merged_value(worksheet, address: str, value: object) -> None:
    worksheet.append({
        "kind": "cell",
        "sheet": "MASTER",
        "address": address,
        "value": "" if value is None else str(value),
        "numberFormat": "@",
    })


def _set_merged_date_text(worksheet, address: str, value: str | date) -> None:
    worksheet.append({
        "kind": "cell",
        "sheet": "MASTER",
        "address": address,
        "value": _as_excel_date_text(value),
        "numberFormat": "@",
    })


def _configure_landscape_a4(page_setup, *, print_area: str | None = None) -> None:
    page_setup.PaperSize = XL_PAPER_A4
    page_setup.Orientation = XL_LANDSCAPE
    page_setup.LeftMargin = MARGIN_POINTS
    page_setup.RightMargin = MARGIN_POINTS
    page_setup.TopMargin = MARGIN_POINTS
    page_setup.BottomMargin = MARGIN_POINTS
    page_setup.HeaderMargin = 0
    page_setup.FooterMargin = 0
    page_setup.CenterHorizontally = True
    page_setup.CenterVertically = False
    page_setup.Zoom = False
    page_setup.FitToPagesWide = 1
    page_setup.FitToPagesTall = False
    if print_area:
        page_setup.PrintArea = print_area


def _apply_excel_alignment(target, alignment: str | None) -> None:
    if alignment == "center":
        target.HorizontalAlignment = XL_CENTER
    elif alignment == "right":
        target.HorizontalAlignment = XL_RIGHT
    elif alignment == "left":
        target.HorizontalAlignment = XL_LEFT


def _apply_excel_cell_action(workbook, action: dict) -> None:
    sheet = workbook.Worksheets(action.get("sheet", "MASTER"))
    target = sheet.Range(action["address"])
    if action.get("fontSize"):
        try:
            target.Font.Size = action["fontSize"]
        except Exception:
            pass
    if action.get("numberFormat"):
        try:
            target.NumberFormat = action["numberFormat"]
        except Exception:
            pass

    value = action.get("value")
    value_type = action.get("valueType", "string")
    if value_type == "formula":
        target.Formula = "" if value is None else str(value)
    elif value_type == "number":
        target.Value = 0 if value in (None, "") else float(value)
    else:
        target.Value = "" if value is None else str(value)

    _apply_excel_alignment(target, action.get("alignment"))


def _apply_excel_page_setup_action(workbook, action: dict) -> None:
    sheet = workbook.Worksheets(action.get("sheet", "MASTER"))
    page_setup = sheet.PageSetup
    page_setup.PaperSize = XL_PAPER_A4
    if action.get("landscape"):
        page_setup.Orientation = XL_LANDSCAPE
    if action.get("printArea"):
        page_setup.PrintArea = action["printArea"]
    if action.get("centerHorizontally") is not None:
        page_setup.CenterHorizontally = bool(action["centerHorizontally"])
    if action.get("centerVertically") is not None:
        page_setup.CenterVertically = bool(action["centerVertically"])
    if action.get("fitToPagesWide") is not None:
        page_setup.Zoom = False
        page_setup.FitToPagesWide = int(action["fitToPagesWide"])
    if action.get("fitToPagesTall") is not None:
        page_setup.Zoom = False
        page_setup.FitToPagesTall = int(action["fitToPagesTall"])
    elif action.get("fitToPagesWide") is not None:
        page_setup.FitToPagesTall = False
    if action.get("marginPoints") is not None:
        margin = float(action["marginPoints"])
        page_setup.LeftMargin = margin
        page_setup.RightMargin = margin
        page_setup.TopMargin = margin
        page_setup.BottomMargin = margin
        page_setup.HeaderMargin = 0
        page_setup.FooterMargin = 0


def _apply_excel_action(workbook, action: dict) -> None:
    if action["kind"] == "cell":
        _apply_excel_cell_action(workbook, action)
    elif action["kind"] == "page_setup":
        _apply_excel_page_setup_action(workbook, action)


def _delete_sheet_if_exists(workbook, sheet_name: str) -> None:
    for sheet in workbook.Worksheets:
        if sheet.Name == sheet_name:
            sheet.Delete()
            return


def _paste_print_range_picture(source_sheet, layout_sheet, source_range: str, *,
                               left: float, top: float, max_width: float,
                               max_height: float) -> None:
    source = source_sheet.Range(source_range)
    source.CopyPicture(Appearance=XL_PRINTER, Format=XL_PICTURE)
    layout_sheet.Paste()
    shape = layout_sheet.Shapes(layout_sheet.Shapes.Count)
    scale = min(max_width / shape.Width, max_height / shape.Height)
    shape.LockAspectRatio = True
    shape.Width = shape.Width * scale
    shape.Left = left + ((max_width - shape.Width) / 2)
    shape.Top = top + ((max_height - shape.Height) / 2)


def _export_kpi_two_pages_per_sheet(workbook, master_sheet, guide_sheet, output_pdf_path: Path) -> None:
    _delete_sheet_if_exists(workbook, KPI_LAYOUT_SHEET_NAME)
    layout = workbook.Worksheets.Add(After=master_sheet)
    layout.Name = KPI_LAYOUT_SHEET_NAME
    layout.Cells.Clear()
    layout.ResetAllPageBreaks()
    source_pages = [
        (master_sheet, source_range) for source_range in KPI_MASTER_PRINT_RANGES
    ]

    for column_index in range(1, 27):
        layout.Columns(column_index).ColumnWidth = 4

    rows_per_output_page = 42
    output_page_count = (len(source_pages) + 1) // 2
    total_rows = output_page_count * rows_per_output_page
    for row_index in range(1, total_rows + 1):
        layout.Rows(row_index).RowHeight = 13.4

    print_area = f"A1:Z{total_rows}"
    _configure_landscape_a4(layout.PageSetup, print_area=print_area)

    layout_width = layout.Range("A1:Z1").Width
    page_height = layout.Range(f"A1:A{rows_per_output_page}").Height
    gutter = 14
    padding = 8
    slot_width = (layout_width - gutter - (padding * 2)) / 2
    slot_height = page_height - (padding * 2)

    for index, (source_sheet, source_range) in enumerate(source_pages):
        output_page_index = index // 2
        slot_index = index % 2
        top = layout.Cells((output_page_index * rows_per_output_page) + 1, 1).Top + padding
        left = layout.Cells((output_page_index * rows_per_output_page) + 1, 1).Left + padding
        if slot_index == 1:
            left += slot_width + gutter

        _paste_print_range_picture(
            source_sheet,
            layout,
            source_range,
            left=left,
            top=top,
            max_width=slot_width,
            max_height=slot_height,
        )

    for page_index in range(1, output_page_count):
        layout.HPageBreaks.Add(Before=layout.Rows((page_index * rows_per_output_page) + 1))

    layout.ExportAsFixedFormat(XL_TYPE_PDF, str(output_pdf_path))


def _run_excel_kpi_export(
    *,
    template_path: Path,
    working_workbook_path: Path,
    output_pdf_path: Path,
    actions: list[dict],
    task_list: str,
    worker_feedback: str = "",
    training_needs: str = "",
    evaluator_feedback: str = "",
) -> None:
    import pythoncom
    import win32com.client

    working_workbook_path.parent.mkdir(parents=True, exist_ok=True)
    output_pdf_path.parent.mkdir(parents=True, exist_ok=True)
    _unprotect_and_prepare_kpi_workbook(
        template_path,
        working_workbook_path,
        task_list,
        worker_feedback=worker_feedback,
        training_needs=training_needs,
        evaluator_feedback=evaluator_feedback,
    )

    pythoncom.CoInitialize()
    excel = None
    workbook = None
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False

        workbook = excel.Workbooks.Open(
            str(working_workbook_path),
            UpdateLinks=0,
            ReadOnly=False,
        )

        for action in actions:
            _apply_excel_action(workbook, action)

        try:
            excel.CalculateFull()
        except Exception:
            workbook.Application.Calculate()

        master_sheet = workbook.Worksheets("MASTER")
        guide_sheet = workbook.Worksheets("Panduan")
        _export_kpi_two_pages_per_sheet(workbook, master_sheet, guide_sheet, output_pdf_path)
        _delete_sheet_if_exists(workbook, KPI_LAYOUT_SHEET_NAME)
        workbook.Save()
    finally:
        if workbook is not None:
            workbook.Close(SaveChanges=False)
        if excel is not None:
            excel.Quit()
        pythoncom.CoUninitialize()


def _kpi_source_page_order(page_count: int) -> list[int]:
    if page_count == 5:
        return [2, 3, 4, 0, 1]
    return list(range(page_count))


def _impose_pdf_two_pages_per_sheet(source_pdf_path: Path, output_pdf_path: Path) -> None:
    from pypdf import PageObject, PdfReader, PdfWriter, Transformation

    reader = PdfReader(str(source_pdf_path))
    writer = PdfWriter()
    if not reader.pages:
        raise RuntimeError(f"KPI source PDF has no pages: {source_pdf_path}")

    first_page = reader.pages[0]
    first_width = float(first_page.mediabox.width)
    first_height = float(first_page.mediabox.height)
    output_width = max(first_width, first_height)
    output_height = min(first_width, first_height)
    slot_width = (output_width - PDF_SLOT_GUTTER_POINTS - (PDF_SLOT_PADDING_POINTS * 2)) / 2
    slot_height = output_height - (PDF_SLOT_PADDING_POINTS * 2)
    raw_pages = len(reader.pages)
    if raw_pages >= 5:
        source_order = [2, 3, 4]
    elif raw_pages >= 3:
        source_order = list(range(raw_pages - 3, raw_pages))
    else:
        source_order = list(range(raw_pages))

    for output_index in range(0, len(source_order), 2):
        output_page = PageObject.create_blank_page(width=output_width, height=output_height)
        for slot_index, source_index in enumerate(source_order[output_index:output_index + 2]):
            if source_index >= len(reader.pages):
                continue
            source_page = copy(reader.pages[source_index])
            source_page.transfer_rotation_to_content()
            source_width = float(source_page.mediabox.width)
            source_height = float(source_page.mediabox.height)
            scale = min(slot_width / source_width, slot_height / source_height)
            left = PDF_SLOT_PADDING_POINTS
            if slot_index == 1:
                left += slot_width + PDF_SLOT_GUTTER_POINTS
            left += (slot_width - (source_width * scale)) / 2
            bottom = PDF_SLOT_PADDING_POINTS + ((slot_height - (source_height * scale)) / 2)
            output_page.merge_transformed_page(
                source_page,
                Transformation().scale(scale).translate(left, bottom),
            )
        writer.add_page(output_page)

    with output_pdf_path.open("wb") as handle:
        writer.write(handle)


def _run_libreoffice_kpi_export(
    *,
    template_path: Path,
    working_workbook_path: Path,
    output_pdf_path: Path,
    actions: list[dict],
    task_list: str,
    worker_feedback: str = "",
    training_needs: str = "",
    evaluator_feedback: str = "",
) -> None:
    _unprotect_and_prepare_kpi_workbook(
        template_path,
        working_workbook_path,
        task_list,
        worker_feedback=worker_feedback,
        training_needs=training_needs,
        evaluator_feedback=evaluator_feedback,
    )
    with tempfile.TemporaryDirectory(dir=output_pdf_path.parent) as temp_dir:
        source_pdf_path = Path(temp_dir) / output_pdf_path.name
        run_calc_pdf_export(
            template_path=template_path,
            working_workbook_path=working_workbook_path,
            output_pdf_path=source_pdf_path,
            actions=actions,
        )
        _impose_pdf_two_pages_per_sheet(source_pdf_path, output_pdf_path)


def generate_kpi_pdf(
    *,
    template_path: str | Path,
    working_workbook_path: str | Path,
    output_pdf_path: str | Path,
    worker_name: str,
    worker_id: str,
    designation: str,
    department: str,
    evaluator_name: str,
    month_label: str,
    task_list: str,
    scores: dict[str, list[int]],
    comments: dict[str, str],
    summary_options: dict[str, str],
    worker_feedback: str,
    training_needs: str,
    evaluator_feedback: str,
    application_date: str,
) -> None:
    template_path = Path(template_path).resolve()
    working_workbook_path = Path(working_workbook_path).resolve()
    output_pdf_path = Path(output_pdf_path).resolve()

    actions: list[dict] = []

    _set_merged_value(actions, "C4", worker_name)
    _set_merged_value(actions, "F4", worker_id)
    _set_merged_value(actions, "C5", designation)
    _set_merged_value(actions, "F5", department)
    _set_merged_value(actions, "C6", evaluator_name)
    _set_merged_value(actions, "F6", month_label)

    for section_key, cell_addresses in KPI_SCORE_CELLS.items():
        section_scores = scores.get(section_key) or []
        for index, cell_address in enumerate(cell_addresses):
            actions.append({
                "kind": "cell",
                "sheet": "MASTER",
                "address": cell_address,
                "value": section_scores[index] if index < len(section_scores) else 0,
                "valueType": "number",
            })

    for section_key, cell_address in KPI_COMMENT_CELLS.items():
        _set_merged_value(actions, cell_address, comments.get(section_key, ""))

    for field_key, cell_address in KPI_OPTION_CELLS.items():
        actions.append({
            "kind": "cell",
            "sheet": "MASTER",
            "address": cell_address,
            "value": summary_options.get(field_key, ""),
            "numberFormat": "@",
        })

    _set_merged_date_text(actions, "G113", application_date)
    actions.append({
        "kind": "page_setup",
        "sheet": "MASTER",
        "printArea": "A1:I124",
        "landscape": True,
        "fitToPagesWide": 1,
        "centerHorizontally": True,
        "marginPoints": MARGIN_POINTS,
    })
    actions.append({
        "kind": "page_setup",
        "sheet": "Panduan",
        "printArea": "A1:E88",
        "landscape": True,
        "fitToPagesWide": 1,
        "centerHorizontally": True,
        "marginPoints": MARGIN_POINTS,
    })

    try:
        _run_excel_kpi_export(
            template_path=template_path,
            working_workbook_path=working_workbook_path,
            output_pdf_path=output_pdf_path,
            actions=actions,
            task_list=task_list,
            worker_feedback=worker_feedback,
            training_needs=training_needs,
            evaluator_feedback=evaluator_feedback,
        )
    except Exception:
        _run_libreoffice_kpi_export(
            template_path=template_path,
            working_workbook_path=working_workbook_path,
            output_pdf_path=output_pdf_path,
            actions=actions,
            task_list=task_list,
            worker_feedback=worker_feedback,
            training_needs=training_needs,
            evaluator_feedback=evaluator_feedback,
        )

    _ensure_kpi_pdf_max_2_pages(output_pdf_path)


def _ensure_kpi_pdf_max_2_pages(output_pdf_path: Path) -> None:
    try:
        from pypdf import PdfReader, PdfWriter
        if not output_pdf_path.is_file():
            return
        reader = PdfReader(str(output_pdf_path))
        if len(reader.pages) > 2:
            writer = PdfWriter()
            for page in reader.pages[:2]:
                writer.add_page(page)
            with output_pdf_path.open("wb") as handle:
                writer.write(handle)
    except Exception:
        pass

